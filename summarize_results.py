import pandas as pd
import os
from pathlib import Path
import re
import openpyxl

def _parse_metrics_from_sheet1(file_path: Path) -> dict[str, float]:
    """
    从 Sheet1 中解析基于文本的指标。
    Sheet1 的格式为 "metric_name: value"。
    """
    metrics = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        if 'Sheet1' not in workbook.sheetnames:
            print(f"    - 错误: 在 {file_path.name} 中未找到 'Sheet1'。")
            return {}

        sheet = workbook['Sheet1']
        
        # 逐行读取内容并用正则表达式匹配
        for row in sheet.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            
            line = str(row[0])
            
            match = re.search(r"faithfulness\s*:\s*(\d+\.?\d*)", line)
            if match:
                metrics['faithfulness'] = float(match.group(1))
                continue

            match = re.search(r"answer_relevancy\s*:\s*(\d+\.?\d*)", line)
            if match:
                metrics['answer_relevancy'] = float(match.group(1))
                continue

            match = re.search(r"context_precision\s*:\s*(\d+\.?\d*)", line)
            if match:
                metrics['context_precision'] = float(match.group(1))
                continue
            
            match = re.search(r"context_recall\s*:\s*(\d+\.?\d*)", line)
            if match:
                metrics['context_recall'] = float(match.group(1))
                continue

        print(f"    - 从 'Sheet1' 中成功解析出 {len(metrics)} 个指标。")
        return metrics

    except Exception as e:
        print(f"    - 错误: 解析 '{file_path.name}' 的 'Sheet1' 时发生错误: {e}")
        return {}


def summarize_evaluation_results():
    """
    汇总所有RAGAS评估结果到一个Excel文件中。
    """
    results_dir = Path("evaluation_results")
    summary_data = []

    # 你提供的文档数量映射
    doc_counts = {
        "covidqa": 4944,
        "delucionqa": 930,
        "expertqa": 7403,
        "hagrid": 6965,
        "msmarco": 21879,
        "techqa": 4066,
        "cuad": 509,
        "emanual": 221,
        "finqa": 7923,
        "hotpotqa": 10506,
        "pubmedqa": 47268,
        "tatqa": 12726,
    }

    # RAGAS指标列名
    METRIC_COLUMNS = ['faithfulness', 'answer_relevancy', 'context_precision', 'context_recall']

    print(f"正在扫描目录: {results_dir.absolute()}")

    # 遍历目录下的所有xlsx文件
    for xlsx_file in results_dir.glob("*.xlsx"):
        filename = xlsx_file.name
        
        # 排除汇总文件自身
        if "summary" in filename.lower():
            continue

        print(f"  -> 正在处理文件: {filename}")
        
        # 从文件名中提取最长匹配的数据集名称
        # 例如，避免 "techqa" 被 "qa" 错误匹配
        matched_keys = [key for key in doc_counts if key in filename]
        if not matched_keys:
            print(f"    - 警告: 无法从文件名 '{filename}' 中提取有效的数据集名称。跳过此文件。")
            continue
            
        dataset_name = max(matched_keys, key=len)

        print(f"    - 识别到数据集: {dataset_name}")

        try:
            # 1. 从 '详细评估结果' sheet 中获取 query 数量
            try:
                details_df = pd.read_excel(xlsx_file, sheet_name='详细评估结果')
                query_count = len(details_df)
                print(f"    - 从 '详细评估结果' sheet 中获取到 Query数: {query_count}")
            except Exception as e:
                print(f"    - 警告: 无法从 '{filename}' 读取 '详细评估结果' sheet 来获取query数: {e}。将query数设为未知。")
                query_count = "未知"

            # 2. 从 'Sheet1' 中直接解析最终指标
            final_metrics = _parse_metrics_from_sheet1(xlsx_file)

            if not final_metrics or len(final_metrics) < 4:
                print(f"    - 警告: 未能从 '{filename}' 的 'Sheet1' 中解析出所有4个指标。跳过此文件。")
                continue
            
            # 获取文档数
            document_count = doc_counts.get(dataset_name, "未知")

            # 准备要添加到汇总列表的数据
            summary_row = {
                "数据集名称": dataset_name,
                "文档数": document_count,
                "query数": query_count,
                **final_metrics
            }

            summary_data.append(summary_row)
            print(f"    - 处理成功: 数据集={dataset_name}, Faithfulness={final_metrics.get('faithfulness', -1):.4f}")

        except Exception as e:
            print(f"    - 错误: 处理文件 '{filename}' 时发生严重错误: {e}")

    if not summary_data:
        print("没有找到或处理任何评估结果文件。")
        return

    # 创建汇总DataFrame
    summary_df = pd.DataFrame(summary_data)
    
    # 按照数据集名称排序
    summary_df = summary_df.sort_values(by="数据集名称").reset_index(drop=True)

    # 定义列顺序
    column_order = ["数据集名称", "文档数", "query数"] + METRIC_COLUMNS
    summary_df = summary_df[column_order]

    # 将汇总结果保存到新的Excel文件
    output_filename = "evaluation_summary.xlsx"
    output_path = results_dir / output_filename
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name="评估结果汇总")
        
        # 自动调整列宽
        worksheet = writer.sheets["评估结果汇总"]
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max(max_length + 2, len(str(column[0].value)) + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"\n汇总完成！结果已保存到: {output_path.absolute()}")
    print(summary_df)

if __name__ == "__main__":
    summarize_evaluation_results()
